#!/usr/bin/env python3
"""
Round-2 validated-data updater.

Reads validated-data/Case 52633_SBMM_SO_202603_Flat_ltw.xlsx and rewrites the
R2 portion of:
  - abp_samples.geojson          (file-of-record)
  - SBMM_ABP_Master_Table_R1_R2.xlsx  (Round 2 sheet)
  - the const DATA = {...} block embedded in index.html
to carry the validated values, the U/J lab qualifiers, the detection limits
for non-detects, sample dates, and the corrected deep depth (2-3 ft).

R1 and EA features are left untouched.
"""
import openpyxl, json, re, copy
from pathlib import Path

FLAT = 'validated-data/Case 52633_SBMM_SO_202603_Flat_ltw.xlsx'
GEOJSON = 'abp_samples.geojson'
MASTER  = 'SBMM_ABP_Master_Table_R1_R2.xlsx'
INDEX   = 'index.html'

ROD = {'Hg':204,'As':6.1,'Sb':51,'Tl':1.3}
CHEM_MAP = {'Mercury':'Hg','Arsenic':'As','Antimony':'Sb','Thallium':'Tl'}
DEPTH_MAP = {'-0.0/0.5':'shallow','-2.0/3.0':'deep','-OM':'OM'}

# ---------- 1) parse validated flat file -------------------------------------
wb = openpyxl.load_workbook(FLAT, data_only=True)
ws = wb['Validation Changes']
hdr = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
H = {h:i+1 for i,h in enumerate(hdr)}
def cell(r,name): return ws.cell(r, H[name]).value

def parse_result(raw, detect, labq):
    """Return (numeric_value_or_None, is_nondetect, detection_limit_or_None)."""
    raw = str(raw).strip()
    nd = (detect == 'N') or raw.startswith('<')
    m  = re.search(r'[-+]?\d*\.?\d+', raw)
    n  = float(m.group()) if m else None
    if nd:
        return None, True, n     # n is the reporting / detection limit
    return n, False, None

val = {}   # val[loc][depth][analyte] = {value, detect, dl, qual, date}
for r in range(2, ws.max_row+1):
    if cell(r,'MATRIX_CODE') != 'SO':
        continue
    name = cell(r,'SAMPLE_NAME') or ''
    m = re.match(r'SBM-ABP-SO-R2-((?:W|E|OS)\d+)(-0\.0/0\.5|-2\.0/3\.0|-OM)(-FD)?$', name)
    if not m: continue
    loc, dsuf, is_fd = m.group(1), m.group(2), bool(m.group(3))
    if is_fd: continue       # primary data only (FDs stay in QC realm)
    chem = CHEM_MAP.get(cell(r,'CHEMICAL_NAME'))
    if not chem: continue
    n, nd, dl = parse_result(cell(r,'REPORT_RESULT_TEXT'), cell(r,'DETECT_FLAG'), cell(r,'LAB_QUALIFIERS'))
    val.setdefault(loc, {}).setdefault(DEPTH_MAP[dsuf], {})[chem] = {
        'value': n, 'detect': not nd, 'dl': dl,
        'qual':  cell(r,'INTERPRETED_QUALIFIERS') or cell(r,'LAB_QUALIFIERS') or '',
        'date':  str(cell(r,'SAMPLE_DATE'))[:10],
    }

print(f"parsed {len(val)} R2 locations from validated flat file")

# ---------- 2) recompute rod_exceed / rod_drivers ----------------------------
def rod_eval(values):
    """values is dict like {Hg:{value,detect,dl}, As:..., ...} for ONE depth.
       Returns ('YES'|'NO'|'NOT SAMPLED', driver_str_or_None)."""
    drivers = []
    any_detect = False
    for a in ('Hg','As','Sb','Tl'):
        v = values.get(a)
        if not v: continue
        any_detect = any_detect or v['detect']
        if v['detect'] and v['value'] is not None and v['value'] > ROD[a]:
            drivers.append(a)
    if not values: return 'NOT SAMPLED', None
    return ('YES' if drivers else 'NO'), (', '.join(drivers) if drivers else None)

# ---------- 3) build new R2 properties dictionaries --------------------------
def fmt(v): return v if v is not None else None  # null in JSON for NDs

new_props = {}   # loc_id -> { the R2 property dict to overwrite }
for loc, depths in val.items():
    sh = depths.get('shallow', {})
    dp = depths.get('deep', {})
    om = depths.get('OM', {})

    # Use the validated-derived exceedance of the SHALLOW horizon as the
    # representative rod_exceed for the sample (same convention as existing data).
    rod_exc, rod_dr = rod_eval(sh)

    p = {
        # core analytes (shallow) — numeric value if detect, null if ND
        'Hg':  fmt(sh.get('Hg',{}).get('value')),
        'As':  fmt(sh.get('As',{}).get('value')),
        'Sb':  fmt(sh.get('Sb',{}).get('value')),
        'Tl':  fmt(sh.get('Tl',{}).get('value')),
        # detection limits for the non-detects (so we can apply 1/2 DL or DL later)
        'Hg_dl': sh.get('Hg',{}).get('dl'),
        'As_dl': sh.get('As',{}).get('dl'),
        'Sb_dl': sh.get('Sb',{}).get('dl'),
        'Tl_dl': sh.get('Tl',{}).get('dl'),
        # lab qualifiers
        'Hg_q': sh.get('Hg',{}).get('qual','') or '',
        'As_q': sh.get('As',{}).get('qual','') or '',
        'Sb_q': sh.get('Sb',{}).get('qual','') or '',
        'Tl_q': sh.get('Tl',{}).get('qual','') or '',
        # deep — present only if deep horizon was sampled
        'Hg_deep': fmt(dp.get('Hg',{}).get('value')) if dp else None,
        'As_deep': fmt(dp.get('As',{}).get('value')) if dp else None,
        'Sb_deep': fmt(dp.get('Sb',{}).get('value')) if dp else None,
        'Tl_deep': fmt(dp.get('Tl',{}).get('value')) if dp else None,
        'Hg_dl_deep': dp.get('Hg',{}).get('dl') if dp else None,
        'As_dl_deep': dp.get('As',{}).get('dl') if dp else None,
        'Sb_dl_deep': dp.get('Sb',{}).get('dl') if dp else None,
        'Tl_dl_deep': dp.get('Tl',{}).get('dl') if dp else None,
        'Hg_q_deep': (dp.get('Hg',{}).get('qual','') or '') if dp else '',
        'As_q_deep': (dp.get('As',{}).get('qual','') or '') if dp else '',
        'Sb_q_deep': (dp.get('Sb',{}).get('qual','') or '') if dp else '',
        'Tl_q_deep': (dp.get('Tl',{}).get('qual','') or '') if dp else '',
        # OM Hg
        'Hg_OM':   fmt(om.get('Hg',{}).get('value')) if om else None,
        'Hg_OM_q': (om.get('Hg',{}).get('qual','') or '') if om else '',
        # exceedance (recomputed from validated detects)
        'rod_exceed':  rod_exc,
        'rod_drivers': rod_dr,
        # sample date (earliest across the three horizons)
        'sample_date': min(filter(None, [
            sh.get('Hg',{}).get('date'),
            dp.get('Hg',{}).get('date'),
            om.get('Hg',{}).get('date'),
        ]), default=None),
        # update depth-string field to reflect 0-0.5 / 2-3 ft naming
        '_depth_intervals': ('shallow + deep' if dp else 'shallow') + (' + OM' if om else ''),
        # validation flag
        'validated': True,
    }
    new_props[loc] = p

# ---------- 4) write updated abp_samples.geojson -----------------------------
gj = json.load(open(GEOJSON))
PRESERVE = {'id','round','sampled','depth','coord_source','field_note','has_fd',
            'not_sampled_reason','notes','OM_effect'}   # keep these from existing R2 props
ANALYTE_FIELDS = {
    'Hg','As','Sb','Tl',
    'Hg_deep','As_deep','Sb_deep','Tl_deep',
    'Hg_OM',
}
QUAL_FIELDS_NEW = {
    'Hg_dl','As_dl','Sb_dl','Tl_dl','Hg_q','As_q','Sb_q','Tl_q',
    'Hg_dl_deep','As_dl_deep','Sb_dl_deep','Tl_dl_deep',
    'Hg_q_deep','As_q_deep','Sb_q_deep','Tl_q_deep',
    'Hg_OM_q','sample_date','validated','_depth_intervals',
}

n_updated = 0
for f in gj['features']:
    p = f['properties']
    if p.get('round') != 2: continue
    if not p.get('sampled'): continue       # skip not-sampled W16/W17/E05/E18
    loc = p['id']
    if loc not in new_props:
        print(f"  WARNING: R2 sample {loc} not found in validated file; left as-is")
        continue
    np = new_props[loc]
    # overwrite analyte fields + add qualifier/DL/date fields
    for k,v in np.items():
        p[k] = v
    # also nudge the human-readable 'depth' string if appropriate (R2 only)
    if np.get('Hg_deep') is not None or any(np.get(a+'_deep') is not None for a in ('Hg','As','Sb','Tl')):
        p['depth'] = 'Shallow + Deep'
    else:
        p['depth'] = 'Shallow'
    n_updated += 1

# bump metadata
gj['metadata']['title']  = 'SBMM ABP Surface Soil Sampling — Rounds 1 & 2 (R2 VALIDATED)'
gj['metadata']['note']   = ('Round 2 analytical data is VALIDATED (Case 52633). '
                            'Several near-detection-limit Sb/Tl/As results were re-qualified as '
                            'non-detect due to equipment-blank contamination (Linda Wilson, '
                            'Jacobs project chemist). Round 1 and EA Historical data are unchanged.')
# (do not touch PMB in metadata since it's no longer used in the UI, but keep present)

json.dump(gj, open(GEOJSON,'w'), indent=2)
print(f"updated {n_updated} R2 features in {GEOJSON}")

# ---------- 5) sync the embedded DATA in index.html --------------------------
html = open(INDEX, encoding='utf-8').read()
# locate const DATA = {...};\n
m = re.search(r'(const DATA = )(\{.*?\});\n', html, flags=re.DOTALL)
assert m, "could not find const DATA in index.html"
minified = json.dumps(gj, separators=(',',':'), ensure_ascii=False)
html = html[:m.start()] + 'const DATA = ' + minified + ';\n' + html[m.end():]
open(INDEX,'w',encoding='utf-8').write(html)
print(f"resynced const DATA in {INDEX}")

# ---------- 6) update Master xlsx Round 2 sheet ------------------------------
mwb = openpyxl.load_workbook(MASTER)
ws = mwb['Round 2 (2026)']
hdr = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
H = {h:i+1 for i,h in enumerate(hdr)}

# add Sample_Date column if not present
if 'Sample_Date' not in H:
    new_col = ws.max_column + 1
    ws.cell(1, new_col, 'Sample_Date')
    H['Sample_Date'] = new_col

def setcell(r, header, value):
    if header in H: ws.cell(r, H[header], value)

ROD_T = {'Hg':204,'As':6.1,'Sb':51,'Tl':1.3}
def exc_text(v_or_None, thr):
    if v_or_None is None: return 'ND'
    return 'YES' if v_or_None > thr else 'NO'

for r in range(2, ws.max_row+1):
    loc = ws.cell(r, H['Location_ID']).value
    if not loc or loc not in new_props: continue
    np = new_props[loc]
    # shallow values + qualifiers
    setcell(r,'Hg_Shallow_mgkg',  np.get('Hg'));  setcell(r,'Hg_Shallow_Q', np.get('Hg_q'))
    setcell(r,'As_Shallow_mgkg',  np.get('As'));  setcell(r,'As_Shallow_Q', np.get('As_q'))
    setcell(r,'Sb_Shallow_mgkg',  np.get('Sb'));  setcell(r,'Sb_Shallow_Q', np.get('Sb_q'))
    setcell(r,'Tl_Shallow_mgkg',  np.get('Tl'));  setcell(r,'Tl_Shallow_Q', np.get('Tl_q'))
    # deep
    setcell(r,'Hg_Deep_mgkg',     np.get('Hg_deep'));   setcell(r,'Hg_Deep_Q', np.get('Hg_q_deep'))
    setcell(r,'As_Deep_mgkg',     np.get('As_deep'));   setcell(r,'As_Deep_Q', np.get('As_q_deep'))
    setcell(r,'Sb_Deep_mgkg',     np.get('Sb_deep'));   setcell(r,'Sb_Deep_Q', np.get('Sb_q_deep'))
    setcell(r,'Tl_Deep_mgkg',     np.get('Tl_deep'));   setcell(r,'Tl_Deep_Q', np.get('Tl_q_deep'))
    # exceedance flags (per analyte) — keep their original column headers
    setcell(r,'Hg_Exc_ROD_204',  exc_text(np.get('Hg'), 204))
    setcell(r,'As_Exc_ROD_6.1',  exc_text(np.get('As'), 6.1))
    setcell(r,'Sb_Exc_ROD_51',   exc_text(np.get('Sb'), 51))
    setcell(r,'Tl_Exc_ROD_1.3',  exc_text(np.get('Tl'), 1.3))
    setcell(r,'Any_ROD_Exc',     np.get('rod_exceed'))
    setcell(r,'ROD_Drivers',     np.get('rod_drivers'))
    setcell(r,'Color_ROD', 'RED (ROD Exc)' if np.get('rod_exceed')=='YES' else 'GREEN (No ROD Exc)')
    # sample date
    setcell(r,'Sample_Date',     np.get('sample_date'))

mwb.save(MASTER)
print(f"updated Round 2 sheet in {MASTER} (added Sample_Date column, refreshed values, added qualifiers)")

# ---------- summary ----------------------------------------------------------
det = {a:0 for a in ROD}
nd  = {a:0 for a in ROD}
for loc,p in new_props.items():
    for a in ROD:
        for suff in ('','_deep'):
            v = p.get(a+suff); dl = p.get(a+'_dl'+suff)
            if v is not None: det[a]+=1
            elif dl is not None: nd[a]+=1
print("\n=== validated R2 detect / ND tallies (shallow + deep) ===")
for a in ROD: print(f"  {a}: {det[a]} detect, {nd[a]} non-detect")
exc = sum(1 for p in new_props.values() if p.get('rod_exceed')=='YES')
print(f"  R2 locations with ROD exceedance (shallow): {exc}")
